from flask import Blueprint, request, render_template, send_file, jsonify
from db import get_db_connection
import psycopg2
from openpyxl import load_workbook
from utils import get_next_id, check_duplicate, validate_links, safe_number, handle_database_error
import os
import pandas as pd
import logging

product_bp = Blueprint('product_bp', __name__)

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

"""
Izveido mapes struktūru augšupielādei.
- Ja mape 'uploads' nepastāv, tā tiek izveidota.
- Nodrošina, ka augšupielādētie faili tiks saglabāti drošā direktorijā.
"""


@product_bp.route('/download_template', methods=['GET'])
def download_template():

    """
    Funkcija, lai lejupielādētu produktu veidni.
    - Apstrādā GET pieprasījumu.
    - Iegūst aktuālo kategoriju sarakstu no datubāzes.
    - Aizpilda Excel veidni ar šīm kategorijām.
    - Atgriež atjauninātu Excel failu lejupielādei.
    """

    connection = get_db_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT kategorija_key, nosaukums FROM kategorijas ORDER BY kategorija_key")
    categories = cursor.fetchall()

    template_path = 'static/products_template.xlsx'

    wb = load_workbook(template_path)

    if 'Category' not in wb.sheetnames:
        wb.create_sheet('Category')

    sheet = wb['Category']
    wb.active = wb.index(sheet) 

    sheet.delete_rows(2, sheet.max_row)

    sheet['A1'] = 'Category (product group)'
    sheet['B1'] = 'Category Key (unique identifier)'

    for index, category in enumerate(categories, start=2):
        sheet[f'A{index}'] = category[1] 
        sheet[f'B{index}'] = category[0]

    wb.save(template_path)

    return send_file(template_path, as_attachment=True)


@product_bp.route('/download_example', methods=['GET'])
def download_example():

    """
    Funkcija, lai lejupielādētu piemēra failu.
    - Apstrādā GET pieprasījumu.
    - Atgriež statisku Excel failu ar piemēru produktu tabulai.
    """

    return send_file('static/products_example.xlsx', as_attachment=True)


@product_bp.route('/upload_products', methods=['POST'])
def upload_products():
    """
    Funkcija produktu augšupielādei no Excel faila.
    - Apstrādā POST pieprasījumu un validē datus.
    - Pievieno jaunus produktus un kategorijas datubāzē.
    """

    connection = get_db_connection()
    cursor = connection.cursor()
    errors = [] 

    if 'file' not in request.files:
        return jsonify({"error": "No file provided."}), 400
    file = request.files['file']

    if file.filename == '':
        return jsonify({"error": "No file selected."}), 400

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        return jsonify({"error": "Invalid file format."}), 400

    next_id = get_next_id(cursor, "produkts")

    for index, row in df.iterrows():
        try:
            nosaukums = str(row['Name']).strip()
            kalorijas = safe_number(row['Energy Value (kcal per kg)'])
            olbaltumvielas = safe_number(row['Proteins (g per kg)'])
            tauki = safe_number(row['Fats (g per kg)'])
            oglhidrati = safe_number(row['Carbohydrates (g per kg)'])
            meris_vieniba = str(row['Unit (measurement unit)']).strip()
            kategorija_key = int(row['Category Key (unique identifier)'])
            kategorija_name = str(row['Category (product group)']).strip()
            vegan = 1 if str(row['Vegan']).strip().lower() == '1' else 0
            saite_maxima = str(row['Maxima Link']).strip() if pd.notna(row['Maxima Link']) else None
            saite_rimi = str(row['Rimi Link']).strip() if pd.notna(row['Rimi Link']) else None

            link_errors = validate_links(
                {"Maxima Link": saite_maxima, "Rimi Link": saite_rimi},
                valid_prefixes=[
                    "https://barbora.lv/",         # Без www
                    "https://www.barbora.lv/",    # С www
                    "https://www.rimi.lv/"
                ]
            )
            if link_errors:
                errors.append(f"Row {index + 1}: " + "; ".join(link_errors))
                continue


            if check_duplicate(cursor, "produkts", "nosaukums", nosaukums):
                errors.append(f"Row {index + 1}: Product name already exists.")
                continue

            if not check_duplicate(cursor, "kategorijas", "kategorija_key", kategorija_key):
                cursor.execute(
                    "INSERT INTO kategorijas (kategorija_key, nosaukums) VALUES (%s, %s)",
                    (kategorija_key, kategorija_name)
                )

            cursor.execute(
                """
                INSERT INTO produkts (
                    id, nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
                    saite_maxima, saite_rimi, meris_vieniba,
                    kategorija_key, vegan
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (next_id, nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
                 saite_maxima, saite_rimi, meris_vieniba, kategorija_key, vegan)
            )
            next_id += 1

        except Exception as e:
            errors.append(f"Row {index + 1}: {str(e)}")
            continue

    connection.commit()

    try:
        os.remove(file_path)
    except Exception as e:
        errors.append(f"File deletion failed: {str(e)}")


    return jsonify({"errors": errors})

@product_bp.route('/manage_products', methods=['GET'])
def manage_products():

    """
    Funkcija, lai pārvaldītu produktus.
    - Apstrādā GET pieprasījumu.
    - Iegūst visu produktu un kategoriju sarakstu no datubāzes.
    - Atgriež HTML veidni ar produktu datiem.
    """

    connection = get_db_connection()
    cursor = connection.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        cursor.execute("""
            SELECT 
                p.id, p.nosaukums, p.kalorijas, p.olbaltumvielas, p.tauki, 
                p.oglhidrati, p.cena_maxima, p.cena_rimi, p.saite_maxima,
                p.saite_rimi, p.meris_vieniba, c.nosaukums AS category_name,
                p.kategorija_key, p.vegan, p.failed_urls
            FROM produkts p
            LEFT JOIN kategorijas c ON p.kategorija_key = c.kategorija_key
            ORDER BY p.nosaukums
        """)
        products = [dict(row) for row in cursor.fetchall()]

        cursor.execute("SELECT kategorija_key, nosaukums FROM kategorijas ORDER BY nosaukums")
        categories = [dict(row) for row in cursor.fetchall()]

        return render_template('manage_products.html', all_products=products, categories=categories)

    except psycopg2.DatabaseError as db_error:
        logging.error(f"Database error: {db_error}")
        return render_template('error.html', error_message="Failed to load products due to a database error.")

    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        return render_template('error.html', error_message="An unexpected error occurred while loading the products.")

@product_bp.route('/update_product', methods=['POST'])
def update_product():

    """
    Funkcija, lai atjauninātu produkta informāciju.
    Apstrādā POST pieprasījumu un atjaunina produkta datus datubāzē.
    """

    connection = get_db_connection()
    cursor = connection.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        product_id = request.form.get('id')
        nosaukums = request.form.get('nosaukums')
        kalorijas = request.form.get('kalorijas') or None
        olbaltumvielas = request.form.get('olbaltumvielas') or None
        tauki = request.form.get('tauki') or None
        oglhidrati = request.form.get('oglhidrati') or None
        cena_maxima = request.form.get('cena_maxima')
        cena_maxima = None if not cena_maxima or cena_maxima == 'null' else float(cena_maxima)

        cena_rimi = request.form.get('cena_rimi')
        cena_rimi = None if not cena_rimi or cena_rimi == 'null' else float(cena_rimi)

        saite_maxima = request.form.get('saite_maxima') or None
        saite_rimi = request.form.get('saite_rimi') or None
        meris_vieniba = request.form.get('meris_vieniba') or None
        kategorija_key = request.form.get('kategorija_key') or None
        vegan = request.form.get('vegan') or None

        cursor.execute("""
            UPDATE produkts SET 
                nosaukums=%s, kalorijas=%s, olbaltumvielas=%s, tauki=%s, oglhidrati=%s,
                cena_maxima=%s, cena_rimi=%s, saite_maxima=%s, saite_rimi=%s,
                meris_vieniba=%s, kategorija_key=%s, vegan=%s
            WHERE id=%s
        """, (
            nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
            cena_maxima, cena_rimi, saite_maxima, saite_rimi,
            meris_vieniba, kategorija_key, vegan, product_id
        ))
        connection.commit()

        return "Product successfully updated!", 200

    except psycopg2.DatabaseError as db_error:
        logging.error(f"Database error: {db_error}")
        return render_template('error.html', error_message="Failed to update product due to a database error.")

    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        return render_template('error.html', error_message="An unexpected error occurred while updating the product.")

@product_bp.route('/get_existing_names', methods=['GET'])
def get_existing_names():
    """
    Funkcija, lai iegūtu visu esošo produktu nosaukumus.
    - Apstrādā GET pieprasījumu.
    - Atgriež produktu nosaukumus mazajiem burtiem JSON formātā.
    """

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        cursor.execute("SELECT LOWER(nosaukums) FROM produkts")
        names = [row[0] for row in cursor.fetchall()]
        return jsonify(names), 200 

    except psycopg2.DatabaseError as db_error:
        logging.error(f"Database error: {db_error}")
        return jsonify({"error": "Failed to retrieve product names."}), 500

    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        return jsonify({"error": "An unexpected error occurred."}), 500

@product_bp.route('/add_product', methods=['POST'])
def add_product():

    """
    Funkcija produktu pievienošanai datubāzei.
    - Apstrādā POST pieprasījumu ar JSON datiem.
    - Validē katra produkta datus un pievieno tos datubāzei, ja tie atbilst prasībām.
    - Atgriež kļūdas vai veiksmīga pievienošanas ziņojumu.
    """

    connection = get_db_connection()
    cursor = connection.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        data = request.get_json()

        if not isinstance(data, list):
            return {"error": "Invalid data format. Expected a list."}, 400

        errors = [] 
        valid_products = []
        next_id = get_next_id(cursor, "produkts")

        for index, product in enumerate(data):
            nosaukums = product.get('nosaukums')
            kalorijas = safe_number(product.get('kalorijas'))
            olbaltumvielas = safe_number(product.get('olbaltumvielas'))
            tauki = safe_number(product.get('tauki'))
            oglhidrati = safe_number(product.get('oglhidrati'))
            cena_maxima = safe_number(product.get('cena_maxima')) or None
            cena_rimi = safe_number(product.get('cena_rimi')) or None
            saite_maxima = product.get('saite_maxima') or None
            saite_rimi = product.get('saite_rimi') or None
            meris_vieniba = product.get('meris_vieniba')
            kategorija_key = product.get('kategorija_key')
            vegan = product.get('vegan')

            if not all([
                nosaukums, 
                kalorijas is not None,
                olbaltumvielas is not None, 
                tauki is not None, 
                oglhidrati is not None, 
                meris_vieniba, 
                kategorija_key
            ]):
                errors.append(f"Row {index + 1}: Required fields are missing.")
                continue

            if check_duplicate(cursor, "produkts", "LOWER(nosaukums)", nosaukums.lower()):
                errors.append(f"Row {index + 1}: A product with this name already exists.")
                continue

            link_errors = validate_links(
                {"Maxima Link": saite_maxima, "Rimi Link": saite_rimi},
                valid_prefixes=[
                    "https://barbora.lv/", 
                    "https://www.barbora.lv/", 
                    "https://www.rimi.lv/"
                ]
            )
            if link_errors:
                errors.append(f"Row {index + 1}: " + "; ".join(link_errors))
                continue

            valid_products.append(
                (
                    next_id, nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
                    cena_maxima, cena_rimi, saite_maxima, saite_rimi,
                    meris_vieniba, kategorija_key, vegan
                )
            )
            next_id += 1

        if errors:
            connection.rollback()
            cursor.close()
            connection.close()
            return {"errors": errors}, 400

        cursor.executemany(
            """
            INSERT INTO produkts (
                id, nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
                cena_maxima, cena_rimi, saite_maxima, saite_rimi,
                meris_vieniba, kategorija_key, vegan
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            valid_products
        )

        connection.commit()
        return {"message": "All products added successfully!"}, 200

    except psycopg2.DatabaseError as db_error:
        connection.rollback()
        logging.error(f"Database error: {db_error}")
        return {"error": "Failed to add products due to a database error."}, 500

    except Exception as e:
        connection.rollback()
        logging.error(f"Unexpected error: {e}")
        return {"error": "An unexpected error occurred while adding products."}, 500

@product_bp.route('/add_single_product', methods=['POST'])
def add_single_product():

    """
    Funkcija, lai pievienotu vienu produktu datubāzei.
    - Apstrādā POST pieprasījumu ar formas datiem.
    - Validē ievadītos datus, pārbauda dublikātus un saites.
    - Pievieno produktu datubāzei, ja visi dati ir pareizi.
    - Atgriež kļūdu paziņojumus vai apstiprinājumu par pievienošanu.
    """

    connection = get_db_connection()
    cursor = connection.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        next_id = get_next_id(cursor, "produkts")

        nosaukums = request.form.get('nosaukums')
        kalorijas = safe_number(request.form.get('kalorijas'))
        olbaltumvielas = safe_number(request.form.get('olbaltumvielas'))
        tauki = safe_number(request.form.get('tauki'))
        oglhidrati = safe_number(request.form.get('oglhidrati'))
        cena_maxima = safe_number(request.form.get('cena_maxima')) or None
        cena_rimi = safe_number(request.form.get('cena_rimi')) or None
        saite_maxima = request.form.get('saite_maxima') or None
        saite_rimi = request.form.get('saite_rimi') or None
        meris_vieniba = request.form.get('meris_vieniba')
        kategorija_key = request.form.get('kategorija_key')
        vegan = request.form.get('vegan')

        if not nosaukums:
            return {"error": "Product name is required."}, 400
        if kalorijas is None or olbaltumvielas is None or tauki is None or oglhidrati is None:
            return {"error": "Nutritional values must be valid."}, 400
        if not meris_vieniba or not kategorija_key:
            return {"error": "Measurement unit and category key are required."}, 400

        if check_duplicate(cursor, "produkts", "nosaukums", nosaukums):
            return {"error": f"Product '{nosaukums}' already exists."}, 400

        link_errors = validate_links(
            {"Maxima Link": saite_maxima, "Rimi Link": saite_rimi},
            valid_prefixes=[
                "https://barbora.lv/", 
                "https://www.barbora.lv/",
                "https://www.rimi.lv/"
            ]
        )
        if link_errors:
            return {"error": "; ".join(link_errors)}, 400


        cursor.execute(
            """
            INSERT INTO produkts (
                id, nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
                cena_maxima, cena_rimi, saite_maxima, saite_rimi,
                meris_vieniba, kategorija_key, vegan
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                next_id, nosaukums, kalorijas, olbaltumvielas, tauki, oglhidrati,
                cena_maxima, cena_rimi, saite_maxima, saite_rimi,
                meris_vieniba, kategorija_key, vegan
            )
        )
        connection.commit()
        return {"message": "Product added successfully!"}, 200

    except psycopg2.DatabaseError as db_error:
        connection.rollback()
        return handle_database_error(db_error)

    except Exception as e:
        connection.rollback()
        logging.error(f"Unexpected error: {e}")
        return {"error": "An unexpected error occurred."}, 500


@product_bp.route('/delete_category/<int:category_id>', methods=['DELETE'])
def delete_category(category_id):

    """
    Funkcija, lai dzēstu kategoriju no datubāzes.
    - Apstrādā DELETE pieprasījumu, izmantojot kategorijas ID.
    - Validē, vai kategorija eksistē, un dzēš to.
    - Atgriež apstiprinājuma ziņojumu vai kļūdas paziņojumu.
    """

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        cursor.execute("SELECT COUNT(*) FROM kategorijas WHERE kategorija_key = %s", (category_id,))
        if cursor.fetchone()[0] == 0:
            return {"error": "Category not found."}, 404

        cursor.execute("DELETE FROM kategorijas WHERE kategorija_key = %s", (category_id,))
        connection.commit()
        return {"message": "Category deleted successfully!"}, 200

    except psycopg2.DatabaseError as db_error:
        connection.rollback()
        return handle_database_error(db_error)

    except Exception as e:
        connection.rollback()
        logging.error(f"Unexpected error: {e}")
        return {"error": "An unexpected error occurred."}, 500

@product_bp.route('/add_category', methods=['POST'])
def add_category():

    """
    Funkcija, lai pievienotu jaunu kategoriju datubāzei.
    - Apstrādā POST pieprasījumu, kas satur kategorijas nosaukumu.
    - Pārbauda, vai nosaukums ir unikāls, un pievieno to datubāzei.
    - Atgriež apstiprinājuma vai kļūdas ziņojumu.
    """

    data = request.get_json()
    category_name = data.get('name', '').strip()

    if not category_name:
        return {"error": "Category name is required."}, 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if check_duplicate(cursor, "kategorijas", "nosaukums", category_name):
            return {"error": "Category already exists."}, 400

        new_key = get_next_id(cursor, "kategorijas", "kategorija_key")

        cursor.execute(
            "INSERT INTO kategorijas (kategorija_key, nosaukums) VALUES (%s, %s)",
            (new_key, category_name)
        )
        connection.commit()
        return {"message": "Category added successfully!", "key": new_key}, 200

    except psycopg2.DatabaseError as db_error:
        connection.rollback()
        return handle_database_error(db_error)

    except Exception as e:
        connection.rollback()
        logging.error(f"Unexpected error: {e}")
        return {"error": "An unexpected error occurred."}, 500

@product_bp.route('/delete_product', methods=['POST'])
def delete_product():

    """
    Funkcija, lai dzēstu produktu no datubāzes.
    Apstrādā POST pieprasījumu, izmantojot produkta ID.
    """

    product_id = request.form.get('product_id')
    if not product_id:
        return "Missing product ID", 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        cursor.execute("DELETE FROM produkts WHERE id = %s", (product_id,))
        connection.commit()
        return "Product deleted", 200
    
    except psycopg2.DatabaseError as db_error:
        logging.error(f"Database error: {db_error}")
        return render_template('error.html', error_message="Failed to delete product due to a database error.")

    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        return render_template('error.html', error_message="An unexpected error occurred while deleting the product.")
